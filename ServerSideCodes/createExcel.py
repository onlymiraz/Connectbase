import sys    
import os
from os.path import exists
import shutil
import logging
import sharedFunctions as s
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, exceptions
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

#--
#-- Global Variables
#--
logDir="D:\\Logs"
loadDir='D:\\DataDump\\LoadLZ'

#-- Table Names
#loadTables = ['SPREAD','ESTIMATE','CIACREVX20']
loadTables = ['APPROVED', 'BUDGETLINE', 'CIACFORFF', 'FFFIELDS', 'FUTUREYEAR', 'SPREAD','GA','ESTIMATE','CIACREVX20','FFAUTHDTLX','FFAUTHDTLZ']

#-- load file names
loadFiles = {
    'APPROVED': ['FFAPROJX'],
    'BUDGETLINE': ['FFBDGTLN20'],
    'CIACFORFF': ['FFCIACOPNX'],
    'FFFIELDS': ['FFFIELDSX'],
    'FUTUREYEAR': ['JMFPRJBGYR'],
    'SPREAD': ['FFSPREAD20'],
    'GA': ['GROSSADDS'],
    'ESTIMATE': ['JFESTPJYR'],
    'FFAUTHDTLX': ['FFAUTHDTLX'],
    'FFAUTHDTLZ': ['FFAUTHDTLZ'],
    'CIACREVX20': ['CIACREVX20']
}

dtypes=[['APPROVED','Proj Num','int'],
        ['APPROVED','Sub Proj Num','int'],
        ['BUDGETLINE','Proj Num', 'int'], 
        ['BUDGETLINE','Sub Proj Num', 'int'], 
        ['BUDGETLINE','Budget Line Num', 'int'],
        ['CIACFORFF','Proj Num', 'int'], 
        ['CIACFORFF','Sub Proj Num', 'int'], 
        ['CIACFORFF','Budget Dollars', 'float'], 
        ['CIACFORFF','Spent Dollars', 'float'],
        ['FFFIELDS','Proj Num', 'int'],
        ['FFFIELDS','Sub Proj Num', 'int'],
        ['FFFIELDS','Class Of Plant', 'str'],
        ['FFFIELDS','Link Code', 'str'],
        ['FFFIELDS','Justification Code', 'int'],
        ['FFFIELDS','Functional Group','str'],
        ['FFFIELDS','Proj Description','str'],
        ['FFFIELDS','Proj Status Code','str'],
        ['FFFIELDS','Approval Code','str'],
        ['FFFIELDS','Proj Type','str'],
        ['FFFIELDS','Billable','str'],
        ['FFFIELDS','Company', 'int'],
        ['FFFIELDS','Exchange Number', 'int'],
        ['FFFIELDS','Operating Area', 'int'],
        ['FFFIELDS','State','str'],
        ['FFFIELDS','Engineer','str'],
        ['FFFIELDS','Project Owner','str'],
        ['FFFIELDS','Approval Date', 'int'],
        ['FFFIELDS','Est StartDate', 'int'],
        ['FFFIELDS','Est Complete Date', 'int'],
        ['FFFIELDS','Actual Start Date', 'int'],
        ['FFFIELDS','Ready For Service Date', 'int'],
        ['FFFIELDS','Tentative Close Date', 'int'],
        ['FFFIELDS','Close Date', 'int'],
        ['FUTUREYEAR','Co', 'int'],
        ['FUTUREYEAR','Area', 'int'],
        ['FUTUREYEAR','Proj#', 'int'],
        ['FUTUREYEAR','SubProj#', 'int'],
        ['FUTUREYEAR','BdgtYr#', 'int'],
        ['FUTUREYEAR','BdgtName', 'str'],
        ['FUTUREYEAR','BdgtLn#', 'int'],
        ['FUTUREYEAR','Main', 'int'],
        ['FUTUREYEAR','Sub', 'int'],
        ['FUTUREYEAR','Release$', 'float'],
        ['SPREAD','Budget Line Number', 'int'],
        ['SPREAD','Budget Line Name', 'str'],
        ['ESTIMATE','Est#','int'],
        ['ESTIMATE','PJ#','int'],
        ['ESTIMATE','Co#','int'],
        ['ESTIMATE','OA','int'],
        ['ESTIMATE','EstStrDte','int'],
        ['ESTIMATE','EstCompDte','int'],
        ['ESTIMATE','APDte','int'],
        ['ESTIMATE','JC','int'],
        ['ESTIMATE','Issue Dte','int'],
        ['ESTIMATE','Sts','str'],
        ['ESTIMATE','COP','str'],
        ['ESTIMATE','Bdgt Yr','int'],
        ['ESTIMATE','ST','str'],
        ['ESTIMATE','Project Description','str'],
        ['ESTIMATE','Link Code','str'],
        ['ESTIMATE','Proj Life (Yrs)','int'],
        ['ESTIMATE','Data Set','int'],
        ['ESTIMATE','Proj Type','str'],
        ['ESTIMATE','FG','str'],
        ['ESTIMATE','Bud/NonBud','str'],
        ['ESTIMATE','AP Code','str'],
        ['CIACREVX20','State','str'],
        ['CIACREVX20','Company Code','int'],
        ['CIACREVX20','PROJ/SUB','int'],
        ['CIACREVX20','STS','str'],
        ['CIACREVX20','Approval Code','str'],
        ['CIACREVX20','BILLABLE','str'],
        ['CIACREVX20','JC','int'],
        ['CIACREVX20','FUNCTION GRP','st'],
        ['CIACREVX20','LINE #','int'],
        ['CIACREVX20','DESC','str'],
        ['CIACREVX20','SO','str'],
        ['CIACREVX20','MAIN','int'],
        ['CIACREVX20','SUB','int'],
        ['CIACREVX20','CC','int'],
        ['CIACREVX20','MO','int'],
        ['CIACREVX20','YR','int'],
        ['CIACREVX20','BUDGET$','float'],
        ['CIACREVX20','SPNT $','float'],
        ['CIACREVX20','PROJ #','int'],
        ['CIACREVX20','SB #','int'],
        ['CIACREVX20','LN CL DT','int'],
        ['CIACREVX20','EXPECTED START DATE','int'],
        ['CIACREVX20','APPROAVL DATE','int'],
        ['CIACREVX20','ADD/CHANGE DATE','int'],
        ['FFAUTHDTLX','Proj Num','int'],
        ['FFAUTHDTLX','Sub Proj Num','int'],
        ['FFAUTHDTLX','Cost Code','int'],
        ['FFAUTHDTLX','','float'],
        ['FFAUTHDTLZ','Proj Num','int'],
        ['FFAUTHDTLZ','Sub Proj Num','int'],
        ['FFAUTHDTLZ','Cost Code','int'],
        ['FFAUTHDTLZ','','float'],
        ['GA','Company','int'],
        ['GA','AcctCode','str'],
        ['GA','Proj','int'],
        ['GA','SubNum','int'],
        ['GA','GAPRJL','int'],
        ['GA','Main','int'],
        ['GA','Sub','int'],
        ['GA','CostCode','int'],
        ['GA','GASORC','str'],
        ['GA','GAP#','str'],
        ['GA','GAPOLN','int'],
        ['GA','GARCPT','str'],
        ['GA','GARCP$','float'],
        ['GA','GALIN$','float'],
        ['GA','GROSS ADDS','float'],
        ['GA','GATDTE','int'],
        ['GA','GAACTY','int'],
        ['GA','GAACTP','int'],
        ['GA','GAL2CD','int'],
        ['GA','GAEDTE','int'],
        ['GA','GAJTCD','int'],
        ['GA','GAOBUD','float'],
        ['GA','GAR1BD','int'],
        ['GA','GAR2BD','int'],
        ['GA','GALDSC','str'],
        ['GA','GA1COD','int'],
        ['GA','GAOPRA','int'],
        ['GA','GAINV#','str'],
        ['GA','GAIDTE','int'],
        ['GA','GAVEND','int'],
        ['GA','GAVNNM','str'],
        ['GA','GACHS2','int'],
        ['GA','GAVOUC','int'],
        ['GA','GAREF','str'],
        ['GA','GADESC','str'],
        ['GA','GAEXCH','str'],
        ['GA','GAPRJT','str'],
        ['GA','GAFUNC','str'],
        ['GA','STATE','str'],
        ['GA','HRS','int'],
        ['GA','QTY','int'],
        ['GA','UNITCOST','float']]


emailTo = ['DL_WAD_Logs@FTR.com']

emailFrom = 'WAD@FTR.com'

#--
#-- Get this script' basename to use for the log file
#--
scriptName = os.path.basename(sys.argv[0])
scriptBase = s.getBasename(scriptName)

#--
#-- Initialize log file
#--
logFile=logDir + "\\" + scriptBase + ".log"
logger=s.initLogger(logFile)

#------------------------------------------------------------------------------
#-- MAIN
#------------------------------------------------------------------------------
s.logIT(logger,"Starting " + scriptName)

#-- Check that local file directory exists
if not os.path.isdir(loadDir):
   s.logIT (logger, "ERROR: " + loadDir + " does not exist")
   s.emailFile(logger,emailTo,emailFrom,emailSub,logFile)
   sys.exit(1)

s.logIT(logger, "------------------------------")

for tableName in loadTables:
   # set the full path to the load file
   s.logIT (logger, tableName + " " + loadFiles[tableName][0])

   if tableName == "FUTUREYEAR" or  \
      tableName == "ESTIMATE" or \
      tableName == 'FFAUTHDTLX' or \
      tableName == 'FFAUTHDTLZ' or \
      tableName == 'CIACREVX20': 
      loadFile = loadDir + "\\" + loadFiles[tableName][0] + ".csv"
      s.logIT (logger, "  Load File: " + loadFile)

      #-- Check if the lz load file exists
      if not os.path.isfile(loadFile):
          s.logIT (logger,"  ERROR: " + loadFile + " does not exist")

      #-- temp File
      tmpFile = loadDir + "\\" + loadFiles[tableName][0] + "_tmp2.csv"
      s.logIT (logger, "  Writing to: " + tmpFile)

      #--
      #-- Read each line in load file, process data and write to tmp file
      #--
      with open(loadFile, encoding='cp437', mode="r", newline='') as icsvFile, open(tmpFile,  encoding='cp437', mode="w", newline='') as tcsvFile:

         csvReader = csv.reader(icsvFile, doublequote=False, delimiter='|', quotechar='"')
         csvWriter = csv.writer(tcsvFile, delimiter='|', doublequote=False, escapechar='\\', quotechar='"', quoting=csv.QUOTE_MINIMAL)

         for row in csvReader:
            nrow = []
            for col in row:
               nrow.append(col)
            csvWriter.writerow(nrow)

      tcsvFile.close()
      icsvFile.close()

      #--
      #-- Replace some special characters in the tmp file and 
      #-- write to output file. Also count line written to output file
      #--
      outFile = loadDir + "\\" + loadFiles[tableName][0] + "_out2.csv"
      s.logIT (logger, "  Writing to: " + outFile)
      fout = open(outFile, encoding="cp437", mode="w")
      try:
          ftmp = open(tmpFile, encoding="cp437", mode="r")
          for line in ftmp:
             # Replace the String based on the pattern
             newline1 = line.replace('\\"','"')
             fout.write(newline1)
      except OSError:
          s.logIT (logger, "  ERROR: Cannot open " + tmpFile)

      fout.close()
      ftmp.close()

   outExcel = loadDir + "\\" + loadFiles[tableName][0] + "_share.xlsx"

   #--
   #-- Get the table column headers and column type
   #--
   s.logIT (logger, "Getting column headers and type for " + outExcel)
   t=0
   colHeader=[]
   colType=[]
   for i in range(len(dtypes)): 
      if dtypes[i][0] == tableName: 
        colHeader.append(dtypes[i][1])
        colType.append(dtypes[i][2])
   colCount = len(colHeader)


   #--
   #-- Initializing spreadsheet with Header row
   #--
   s.logIT (logger, "Inserting Header row in " + outExcel)

   wb = Workbook(write_only=True)
   ws = wb.create_sheet()

   ws.freeze_panes = "A2"
   ws.print_title_rows:'1:1'

   cellArray=[]
   for h in colHeader:
      cell = WriteOnlyCell(ws,h)
      cell.font = Font(bold=True,color="FFFFFF")
      cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type = "solid")
      cellArray.append(cell)

   ws.append([c for c in cellArray])

   #--
   #-- Loading CSV data into Excel spreadsheet
   #--
   s.logIT (logger, "Loading data into " + outExcel)

   if tableName == "FUTUREYEAR" or \
      tableName == "ESTIMATE" or \
      tableName == 'FFAUTHDTLX' or \
      tableName == 'FFAUTHDTLZ' or \
      tableName == 'CIACREVX20': 
      inCSV = loadDir + "\\" + loadFiles[tableName][0] + "_out2.csv"
   else:
      inCSV = loadDir + "\\" + loadFiles[tableName][0] + "_out.csv"
   r=1
   thin = Side(border_style="thin", color="70AD47")
   with open(inCSV, encoding='cp437', mode="r", newline='') as f:
      reader = csv.reader(f, delimiter='|')
      for row in reader:
         cellArray=[]
         c=0
         for col in row:
            col = ILLEGAL_CHARACTERS_RE.sub('', col)

            if colType[c] == "int":
               try:
                  col = int(col)
               except ValueError:
                  col = str(col)

            if colType[c] == "float":
               try:
                  col = float(col)
               except ValueError:
                  col = str(col)

            cell = WriteOnlyCell(ws,col)

            if r % 2 !=0:
               cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type = "solid")
               cell.border = Border(top=thin, bottom=thin)

            cellArray.append(cell)

            c+=1

         ws.append([a for a in cellArray])
          
         r+=1

         #if r % 10000 == 0:
         #   s.logIT (logger, "Loaded " + str(r) + " rows into xlsx")
         #   s.logIT (logger, "*** BREAKING OUT FOR TESTING ***")
         #   break;

         if r % 10000 == 0:
            s.logIT (logger, "Loaded " + str(r) + " rows into xlsx")


   wb.save(outExcel)

   wb.close()

   #--
   #-- Additional Formatting
   #--
   s.logIT (logger, "Formatting " + outExcel)
   wb = load_workbook(outExcel)
   ws = wb.active

   for i in range(1, ws.max_column+1):
      colLetter=get_column_letter(i)
      ws.column_dimensions[get_column_letter(i)].bestFit = True
      ws.column_dimensions[get_column_letter(i)].auto_size = True
      colWidth=ws.column_dimensions[get_column_letter(i)].width
      ws.column_dimensions[get_column_letter(i)].width = colWidth + 5

   ws.auto_filter.ref = ws.dimensions

   wb.save(outExcel)

   #if os.path.isfile(tmpFile):
   #   os.unlink(tmpFile)

   s.logIT (logger, " ")

#--
#-- Look for ERROR in the log file
#--
if s.stringInFile(logger,"ERROR",logFile):
   emailSub="[ ERROR ] " + scriptName 
else:
   emailSub="[ SUCCESS ] " + scriptName 

s.emailFile(logger,emailTo,emailFrom,emailSub,logFile)
   
s.logIT(logger,"Finished " + scriptName)
